#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
结果导出模块 - Excel报表生成

功能：
1. 整合下载和分析结果
2. 生成详细的Excel报表
3. 支持多种导出格式
4. 提供统计分析功能
"""

import os
import pandas as pd
from typing import List, Dict, Optional
import json
from datetime import datetime
import numpy as np
import tempfile
import shutil
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, Reference
from openpyxl.utils.dataframe import dataframe_to_rows
import matplotlib.pyplot as plt
import matplotlib.patches as patches
from PIL import Image, ImageDraw, ImageFont
import cv2

class ResultExporter:
    """结果导出器"""
    
    def __init__(self):
        """初始化导出器"""
        self.results_data = []
        self.summary_stats = {}
    
    def _load_image_with_chinese_path(self, image_path: str) -> Optional[np.ndarray]:
        """
        加载图片，支持中文路径
        
        Args:
            image_path: 图片路径
            
        Returns:
            图片数组（BGR格式）或None
        """
        try:
            # 方法1：直接使用cv2.imread
            img = cv2.imread(image_path)
            if img is not None:
                return img
        except:
            pass
        
        try:
            # 方法2：使用PIL读取后转换为cv2格式
            from PIL import Image
            pil_img = Image.open(image_path).convert('RGB')
            # PIL是RGB格式，cv2是BGR格式，需要转换
            img_array = np.array(pil_img)
            img_bgr = cv2.cvtColor(img_array, cv2.COLOR_RGB2BGR)
            return img_bgr
        except:
            pass
        
        try:
            # 方法3：使用numpy读取字节流
            with open(image_path, 'rb') as f:
                img_data = f.read()
            img_array = np.frombuffer(img_data, np.uint8)
            img = cv2.imdecode(img_array, cv2.IMREAD_COLOR)
            if img is not None:
                return img
        except:
            pass
        
        print(f"无法加载图片: {image_path}")
        return None
    
    def add_result(self, download_result: Dict, analysis_result: Dict):
        """
        添加单个结果记录
        
        Args:
            download_result: 下载结果
            analysis_result: 分析结果
        """
        # 合并下载和分析结果
        combined_result = {
            # 基本信息
            'longitude': download_result.get('lng', 0),
            'latitude': download_result.get('lat', 0),
            'download_success': download_result.get('success', False),
            'download_time': download_result.get('download_time', ''),
            
            # 文件路径
            'original_image_path': download_result.get('filepath', ''),
            'comprehensive_analysis_path': self._get_comprehensive_analysis_path(download_result.get('filepath', ''), analysis_result),
            
            # 分析结果
            'green_view_rate': analysis_result.get('green_view_rate', 0.0),
            'vegetation_pixels': analysis_result.get('vegetation_pixels', 0),
            'total_pixels': analysis_result.get('total_pixels', 0),
            
            # 错误信息
            'download_error': download_result.get('error', ''),
            'analysis_error': analysis_result.get('error', ''),
            
            # 文件大小
            'file_size_bytes': download_result.get('file_size', 0)
        }
        
        # 添加类别分布信息
        class_distribution = analysis_result.get('class_distribution', {})
        for class_name, class_info in class_distribution.items():
            combined_result[f'{class_name}_percentage'] = class_info.get('percentage', 0.0)
            combined_result[f'{class_name}_pixels'] = class_info.get('pixels', 0)
        
        self.results_data.append(combined_result)
    
    def _get_comprehensive_analysis_path(self, original_image_path: str, analysis_result: Dict) -> str:
        """
        获取综合分析图片路径
        
        Args:
            original_image_path: 原始图片路径
            analysis_result: 分析结果
            
        Returns:
            综合分析图片路径
        """
        if not original_image_path:
            return ''
            
        try:
            # 获取原图文件名（不含扩展名）
            filename = os.path.splitext(os.path.basename(original_image_path))[0]
            
            # 获取原图所在目录的父目录
            parent_dir = os.path.dirname(os.path.dirname(original_image_path))
            if not parent_dir:
                parent_dir = 'output'
                
            # 构建综合分析图片路径
            comprehensive_dir = os.path.join(parent_dir, 'comprehensive_analysis')
            comprehensive_path = os.path.join(comprehensive_dir, f"{filename}_comprehensive_analysis.png")
            
            return comprehensive_path
            
        except Exception as e:
            print(f"生成综合分析图片路径失败: {e}")
            return ''
    
    def add_local_image_result(self, analysis_result: Dict):
        """
        添加本地图片分析结果记录
        
        Args:
            analysis_result: 分析结果
        """
        # 从分析结果中提取图片路径信息
        image_path = analysis_result.get('image_path', analysis_result.get('original_image_path', ''))
        filename = os.path.basename(image_path) if image_path else ''
        
        combined_result = {
            # 基本信息（本地图片没有经纬度）
            'filename': filename,
            'image_path': image_path,
            
            # 文件路径
            'original_image_path': image_path,
            'comprehensive_analysis_path': self._get_comprehensive_analysis_path(image_path, analysis_result),
            
            # 分析结果
            'green_view_rate': analysis_result.get('green_view_rate', 0.0),
            'vegetation_pixels': analysis_result.get('vegetation_pixels', 0),
            'total_pixels': analysis_result.get('total_pixels', 0),
            
            # 错误信息
            'analysis_error': analysis_result.get('error', ''),
            
            # 分析时间
            'analysis_time': analysis_result.get('analysis_time', datetime.now().strftime('%Y-%m-%d %H:%M:%S'))
        }
        
        # 添加类别分布信息
        class_distribution = analysis_result.get('class_distribution', {})
        for class_name, class_info in class_distribution.items():
            combined_result[f'{class_name}_percentage'] = class_info.get('percentage', 0.0)
            combined_result[f'{class_name}_pixels'] = class_info.get('pixels', 0)
        
        self.results_data.append(combined_result)
    
    def add_batch_results(self, download_results: List[Dict], analysis_results: List[Dict]):
        """
        批量添加结果记录
        
        Args:
            download_results: 下载结果列表
            analysis_results: 分析结果列表
        """
        # 确保两个列表长度一致
        min_length = min(len(download_results), len(analysis_results))
        
        for i in range(min_length):
            self.add_result(download_results[i], analysis_results[i])
    
    def add_batch_local_results(self, analysis_results: List[Dict]):
        """
        批量添加本地图片分析结果
        
        Args:
            analysis_results: 分析结果列表
        """
        for analysis_result in analysis_results:
            self.add_local_image_result(analysis_result)
    
    def calculate_summary_statistics(self) -> Dict:
        """
        计算汇总统计信息
        
        Returns:
            统计信息字典
        """
        if not self.results_data:
            return {}
        
        df = pd.DataFrame(self.results_data)
        
        # 基本统计
        total_images = len(df)
        
        # 检查是否有download_success字段（街景模式）
        if 'download_success' in df.columns:
            successful_downloads = df['download_success'].sum()
            download_success_rate = (successful_downloads / total_images * 100) if total_images > 0 else 0
        else:
            # 本地图片模式，所有图片都算作成功下载
            successful_downloads = total_images
            download_success_rate = 100.0
        
        successful_analyses = df[df['green_view_rate'] > 0].shape[0]
        
        # 绿视率统计
        green_rates = df[df['green_view_rate'] > 0]['green_view_rate']
        
        stats = {
            'total_images': total_images,
            'successful_downloads': int(successful_downloads),
            'successful_analyses': successful_analyses,
            'download_success_rate': download_success_rate,
            'analysis_success_rate': (successful_analyses / total_images * 100) if total_images > 0 else 0,
        }
        
        if len(green_rates) > 0:
            stats.update({
                'green_view_rate_mean': float(green_rates.mean()),
                'green_view_rate_median': float(green_rates.median()),
                'green_view_rate_std': float(green_rates.std()),
                'green_view_rate_min': float(green_rates.min()),
                'green_view_rate_max': float(green_rates.max()),
                'green_view_rate_q25': float(green_rates.quantile(0.25)),
                'green_view_rate_q75': float(green_rates.quantile(0.75))
            })
        
        # 绿视率分级统计
        if len(green_rates) > 0:
            stats['green_view_distribution'] = {
                'very_low (0-10%)': int(((green_rates >= 0) & (green_rates < 10)).sum()),
                'low (10-20%)': int(((green_rates >= 10) & (green_rates < 20)).sum()),
                'medium (20-30%)': int(((green_rates >= 20) & (green_rates < 30)).sum()),
                'high (30-40%)': int(((green_rates >= 30) & (green_rates < 40)).sum()),
                'very_high (40%+)': int((green_rates >= 40).sum())
            }
        
        self.summary_stats = stats
        return stats
    
    def export_to_excel(self, output_path: str, include_charts: bool = True) -> bool:
        """
        导出结果到Excel文件
        
        Args:
            output_path: 输出文件路径
            include_charts: 是否包含图表
            
        Returns:
            是否导出成功
        """
        try:
            if not self.results_data:
                print("没有数据可导出")
                return False
            
            # 创建Excel工作簿
            wb = Workbook()
            
            # 删除默认工作表
            wb.remove(wb.active)
            
            # 创建详细结果工作表
            self._create_detailed_results_sheet(wb)
            
            # 创建统计汇总工作表
            self._create_summary_sheet(wb)
            
            # 创建图表工作表
            if include_charts:
                self._create_charts_sheet(wb)
            
            # 保存文件
            wb.save(output_path)
            print(f"Excel报表已保存到: {output_path}")
            return True
            
        except Exception as e:
            import traceback
            error_details = traceback.format_exc()
            print(f"导出Excel失败: {e}")
            print(f"详细错误信息: {error_details}")
            return False
    
    def _create_detailed_results_sheet(self, wb: Workbook):
        """
        创建详细结果工作表
        
        Args:
            wb: Excel工作簿
        """
        ws = wb.create_sheet("详细结果")
        
        # 创建DataFrame
        df = pd.DataFrame(self.results_data)
        
        # 检查数据类型（是否包含经纬度信息）
        has_coordinates = 'longitude' in df.columns and 'latitude' in df.columns
        has_filename = 'filename' in df.columns
        
        if has_coordinates:
            # 街景模式：包含经纬度信息
            column_order = [
                'longitude', 'latitude', 'green_view_rate', 'vegetation_pixels', 'total_pixels',
                'original_image_path', 'vegetation_highlight_path', 'segmentation_overlay_path',
                'download_success', 'download_time', 'file_size_bytes',
                'download_error', 'analysis_error'
            ]
        else:
            # 本地图片模式：包含文件名信息
            column_order = [
                'filename', 'green_view_rate', 'vegetation_pixels', 'total_pixels',
                'original_image_path', 'vegetation_highlight_path', 'segmentation_overlay_path',
                'analysis_time', 'analysis_error'
            ]
        
        # 添加其他列（类别分布等）
        other_columns = [col for col in df.columns if col not in column_order]
        column_order.extend(other_columns)
        
        # 重新排列DataFrame
        df = df.reindex(columns=[col for col in column_order if col in df.columns])
        
        # 写入数据
        for r in dataframe_to_rows(df, index=False, header=True):
            ws.append(r)
        
        # 设置样式
        self._apply_table_style(ws, len(df) + 1)
    
    def _create_summary_sheet(self, wb: Workbook):
        """
        创建统计汇总工作表
        
        Args:
            wb: Excel工作簿
        """
        ws = wb.create_sheet("统计汇总")
        
        # 计算统计信息
        stats = self.calculate_summary_statistics()
        
        # 写入基本统计信息
        ws.append(["项目", "数值", "单位"])
        ws.append(["总图片数", stats.get('total_images', 0), "张"])
        ws.append(["下载成功数", stats.get('successful_downloads', 0), "张"])
        ws.append(["分析成功数", stats.get('successful_analyses', 0), "张"])
        ws.append(["下载成功率", f"{stats.get('download_success_rate', 0):.2f}", "%"])
        ws.append(["分析成功率", f"{stats.get('analysis_success_rate', 0):.2f}", "%"])
        
        # 添加空行
        ws.append([])
        
        # 绿视率统计
        if 'green_view_rate_mean' in stats:
            ws.append(["绿视率统计", "", ""])
            ws.append(["平均值", f"{stats['green_view_rate_mean']:.2f}", "%"])
            ws.append(["中位数", f"{stats['green_view_rate_median']:.2f}", "%"])
            ws.append(["标准差", f"{stats['green_view_rate_std']:.2f}", "%"])
            ws.append(["最小值", f"{stats['green_view_rate_min']:.2f}", "%"])
            ws.append(["最大值", f"{stats['green_view_rate_max']:.2f}", "%"])
            ws.append(["25%分位数", f"{stats['green_view_rate_q25']:.2f}", "%"])
            ws.append(["75%分位数", f"{stats['green_view_rate_q75']:.2f}", "%"])
        
        # 添加空行
        ws.append([])
        
        # 绿视率分布
        if 'green_view_distribution' in stats:
            ws.append(["绿视率分布", "图片数量", "占比"])
            total_analyzed = stats.get('successful_analyses', 1)
            for level, count in stats['green_view_distribution'].items():
                percentage = (count / total_analyzed * 100) if total_analyzed > 0 else 0
                ws.append([level, count, f"{percentage:.1f}%"])
        
        # 设置样式
        self._apply_summary_style(ws)
    
    def _create_charts_sheet(self, wb: Workbook):
        """
        创建图表工作表
        
        Args:
            wb: Excel工作簿
        """
        ws = wb.create_sheet("图表分析")
        
        # 这里可以添加图表创建逻辑
        # 由于openpyxl的图表功能相对复杂，这里先添加基本框架
        ws.append(["图表分析"])
        ws.append(["绿视率分布图表将在此显示"])
    
    def _apply_table_style(self, ws, num_rows: int):
        """
        应用表格样式
        
        Args:
            ws: 工作表
            num_rows: 行数
        """
        # 设置标题行样式
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
        
        # 设置边框
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for row in ws.iter_rows(min_row=1, max_row=num_rows):
            for cell in row:
                cell.border = thin_border
        
        # 自动调整列宽
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    def _apply_summary_style(self, ws):
        """
        应用汇总表样式
        
        Args:
            ws: 工作表
        """
        # 设置标题样式
        title_font = Font(bold=True, size=12)
        for row in ws.iter_rows():
            if row[0].value in ["项目", "绿视率统计", "绿视率分布"]:
                for cell in row:
                    cell.font = title_font
    
    def export_to_csv(self, output_path: str) -> bool:
        """
        导出结果到CSV文件
        
        Args:
            output_path: 输出文件路径
            
        Returns:
            是否导出成功
        """
        try:
            if not self.results_data:
                print("没有数据可导出")
                return False
            
            df = pd.DataFrame(self.results_data)
            df.to_csv(output_path, index=False, encoding='utf-8-sig')
            print(f"CSV文件已保存到: {output_path}")
            return True
            
        except Exception as e:
            print(f"导出CSV失败: {e}")
            return False
    
    def export_summary_json(self, output_path: str) -> bool:
        """
        导出统计汇总到JSON文件
        
        Args:
            output_path: 输出文件路径
            
        Returns:
            是否导出成功
        """
        try:
            stats = self.calculate_summary_statistics()
            
            # 添加导出时间
            stats['export_time'] = datetime.now().isoformat()
            
            with open(output_path, 'w', encoding='utf-8') as f:
                json.dump(stats, f, ensure_ascii=False, indent=2)
            
            print(f"统计汇总已保存到: {output_path}")
            return True
            
        except Exception as e:
            print(f"导出JSON失败: {e}")
            return False
    
    def generate_comprehensive_analysis_image(self, analysis_result: Dict, output_path: str) -> bool:
        """
        生成综合分析图片，包含各类别掩膜图和类别占比图的单张合并图片
        
        Args:
            analysis_result: 分析结果字典
            output_path: 输出图片路径
            
        Returns:
            bool: 是否成功生成
        """
        try:
            # 设置中文字体和样式
            plt.rcParams['font.sans-serif'] = ['SimHei', 'Microsoft YaHei', 'Arial Unicode MS']
            plt.rcParams['axes.unicode_minus'] = False
            plt.rcParams['figure.facecolor'] = 'white'
            plt.rcParams['axes.facecolor'] = 'white'
            
            # 创建图形 - 1行2列布局，优化尺寸和间距
            fig, (ax1, ax2) = plt.subplots(1, 2, figsize=(16, 8))
            fig.suptitle('图像分割分析结果', 
                        fontsize=18, fontweight='bold', y=0.95, color='#2c3e50')
            
            # 调整子图间距
            plt.subplots_adjust(left=0.05, right=0.95, top=0.88, bottom=0.1, wspace=0.15)
            
            # 左侧：各类别掩膜图
            segmentation_data = analysis_result.get('segmentation_map')
            if segmentation_data is not None:
                print("正在生成掩膜可视化...")
                seg_img = self._create_enhanced_segmentation_visualization(segmentation_data)
                ax1.imshow(seg_img)
                ax1.set_title('各类别语义分割掩膜', fontsize=14, fontweight='bold')
                ax1.axis('off')
            else:
                ax1.text(0.5, 0.5, '无分割数据', ha='center', va='center',
                        transform=ax1.transAxes, fontsize=12)
                ax1.set_title('各类别语义分割掩膜', fontsize=14, fontweight='bold')
                ax1.axis('off')
            
            # 右侧：类别占比统计图
            class_distribution = analysis_result.get('class_distribution', {})
            print(f"类别分布数据: {class_distribution}")
            
            if class_distribution:
                # 收集所有有效类别数据
                valid_classes = []
                for class_name, class_info in class_distribution.items():
                    percentage = class_info.get('percentage', 0.0)
                    if percentage > 0.1:  # 显示占比大于0.1%的类别
                        valid_classes.append((class_name, percentage))
                
                if valid_classes:
                    # 按百分比排序
                    valid_classes.sort(key=lambda x: x[1], reverse=True)
                    
                    # 分离类别名称和百分比
                    class_names = [item[0] for item in valid_classes]
                    percentages = [item[1] for item in valid_classes]
                    
                    print(f"有效类别: {class_names}")
                    print(f"对应百分比: {percentages}")
                    
                    # 检查是否有超过1%的类别
                    high_percentage_classes = [p for p in percentages if p > 1.0]
                    
                    # 使用水平条形图显示所有有效类别，更直观简洁
                    # 限制显示前10个类别以避免图表过于拥挤
                    display_classes = valid_classes[:10]
                    names = [item[0] for item in display_classes]
                    values = [item[1] for item in display_classes]
                    
                    # 定义与掩膜一致的颜色映射
                    class_color_map = {
                        'road': '#696969',        # 深灰色
                        'sidewalk': '#8B4513',    # 棕色
                        'building': '#464646',    # 深灰色
                        'wall': '#808080',        # 灰色
                        'fence': '#A0522D',       # 棕褐色
                        'pole': '#A9A9A9',        # 浅灰色
                        'traffic_light': '#FFD700', # 金色
                        'traffic_sign': '#FFA500',  # 橙色
                        'vegetation': '#228B22',    # 森林绿
                        'terrain': '#90EE90',       # 浅绿色
                        'sky': '#87CEEB',          # 天蓝色
                        'person': '#DC143C',       # 深红色
                        'rider': '#FF4500',        # 橙红色
                        'car': '#00008B',          # 深蓝色
                        'truck': '#191970',        # 午夜蓝
                        'bus': '#006400',          # 深绿色
                        'train': '#483D8B',        # 深紫色
                        'motorcycle': '#8A2BE2',   # 蓝紫色
                        'bicycle': '#FF1493'       # 深粉色
                    }
                    
                    # 为每个类别分配对应的颜色
                    colors = []
                    for name in names:
                        if name in class_color_map:
                            colors.append(class_color_map[name])
                        else:
                            colors.append('#808080')  # 默认灰色
                    
                    y_pos = np.arange(len(names))
                    bars = ax2.barh(y_pos, values, color=colors, alpha=0.8, edgecolor='white', linewidth=1)
                    
                    # 设置标签和标题
                    ax2.set_yticks(y_pos)
                    ax2.set_yticklabels(names, fontsize=10, fontweight='bold')
                    ax2.set_xlabel('占比 (%)', fontsize=12, fontweight='bold')
                    ax2.set_title('类别分布统计', fontsize=14, fontweight='bold')
                    
                    # 在条形图上显示数值
                    for i, (bar, value) in enumerate(zip(bars, values)):
                        ax2.text(bar.get_width() + max(values)*0.01, 
                               bar.get_y() + bar.get_height()/2,
                               f'{value:.1f}%', ha='left', va='center', 
                               fontsize=9, fontweight='bold', color='#333333')
                    
                    # 设置网格和样式
                    ax2.grid(True, axis='x', alpha=0.3, linestyle='--')
                    ax2.set_axisbelow(True)
                    
                    # 调整布局，确保标签不被截断
                    ax2.set_xlim(0, max(values) * 1.15)
                    
                    # 反转y轴，使最大值在顶部
                    ax2.invert_yaxis()
                else:
                    ax2.text(0.5, 0.5, '无有效类别数据\n(所有类别占比<0.1%)', ha='center', va='center',
                           transform=ax2.transAxes, fontsize=12)
                    ax2.set_title('分割结果分布', fontsize=14, fontweight='bold')
                    ax2.axis('off')
            else:
                ax2.text(0.5, 0.5, '无类别分布数据', ha='center', va='center',
                       transform=ax2.transAxes, fontsize=12)
                ax2.set_title('分割结果分布', fontsize=14, fontweight='bold')
                ax2.axis('off')
            
            # 最终布局调整
            plt.tight_layout()
            
            # 确保输出目录存在
            os.makedirs(os.path.dirname(output_path), exist_ok=True)
            
            # 保存图片 - 支持中文路径
            try:
                plt.savefig(output_path, dpi=300, bbox_inches='tight', 
                           facecolor='white', edgecolor='none')
                print(f"综合分析图片已保存到: {output_path}")
            except (OSError, UnicodeEncodeError) as e:
                # 如果直接保存失败，尝试使用临时文件名保存后重命名
                temp_path = tempfile.mktemp(suffix='.png')
                plt.savefig(temp_path, dpi=300, bbox_inches='tight', 
                           facecolor='white', edgecolor='none')
                shutil.move(temp_path, output_path)
                print(f"综合分析图片已保存到: {output_path} (通过临时文件)")
            
            plt.close()
            return True
            
        except Exception as e:
            print(f"生成综合分析图片失败: {e}")
            import traceback
            traceback.print_exc()
            if 'fig' in locals():
                plt.close()
            return False
    
    def _create_segmentation_visualization(self, segmentation_map: np.ndarray) -> np.ndarray:
        """
        创建彩色分割可视化图像
        
        Args:
            segmentation_map: 分割结果图
            
        Returns:
            彩色分割图像
        """
        # 优化的颜色映射，使用更协调的配色方案
        color_map = {
            0: [105, 105, 105],  # road - 深灰色
            1: [139, 69, 19],    # sidewalk - 棕色
            2: [70, 70, 70],     # building - 深灰色
            3: [128, 128, 128],  # wall - 灰色
            4: [160, 82, 45],    # fence - 棕褐色
            5: [169, 169, 169],  # pole - 浅灰色
            6: [255, 215, 0],    # traffic_light - 金色
            7: [255, 165, 0],    # traffic_sign - 橙色
            8: [34, 139, 34],    # vegetation - 森林绿
            9: [144, 238, 144],  # terrain - 浅绿色
            10: [135, 206, 235], # sky - 天蓝色
            11: [220, 20, 60],   # person - 深红色
            12: [255, 69, 0],    # rider - 橙红色
            13: [0, 0, 139],     # car - 深蓝色
            14: [25, 25, 112],   # truck - 午夜蓝
            15: [0, 100, 0],     # bus - 深绿色
            16: [72, 61, 139],   # train - 深紫色
            17: [138, 43, 226],  # motorcycle - 蓝紫色
            18: [255, 20, 147]   # bicycle - 深粉色
        }
        
        # 创建彩色图像
        h, w = segmentation_map.shape
        color_image = np.zeros((h, w, 3), dtype=np.uint8)
        
        for class_id, color in color_map.items():
            mask = segmentation_map == class_id
            color_image[mask] = color
            
        return color_image
    
    def _create_enhanced_segmentation_visualization(self, segmentation_map: np.ndarray) -> np.ndarray:
        """
        创建增强的分割可视化图像，确保不同类别有明显的颜色区分
        
        Args:
            segmentation_map: 分割结果图
            
        Returns:
            增强的彩色分割图像
        """
        # 使用与基础颜色映射一致的增强配色方案
        enhanced_color_map = {
            0: [105, 105, 105],  # road - 深灰色
            1: [139, 69, 19],    # sidewalk - 棕色
            2: [70, 70, 70],     # building - 深灰色
            3: [128, 128, 128],  # wall - 灰色
            4: [160, 82, 45],    # fence - 棕褐色
            5: [169, 169, 169],  # pole - 浅灰色
            6: [255, 215, 0],    # traffic_light - 金色
            7: [255, 165, 0],    # traffic_sign - 橙色
            8: [34, 139, 34],    # vegetation - 森林绿
            9: [144, 238, 144],  # terrain - 浅绿色
            10: [135, 206, 235], # sky - 天蓝色
            11: [220, 20, 60],   # person - 深红色
            12: [255, 69, 0],    # rider - 橙红色
            13: [0, 0, 139],     # car - 深蓝色
            14: [25, 25, 112],   # truck - 午夜蓝
            15: [0, 100, 0],     # bus - 深绿色
            16: [72, 61, 139],   # train - 深紫色
            17: [138, 43, 226],  # motorcycle - 蓝紫色
            18: [255, 20, 147]   # bicycle - 深粉色
        }
        
        # 创建彩色图像
        h, w = segmentation_map.shape
        color_image = np.zeros((h, w, 3), dtype=np.uint8)
        
        # 获取图像中实际存在的类别
        unique_classes = np.unique(segmentation_map)
        
        for class_id in unique_classes:
            if class_id in enhanced_color_map:
                mask = segmentation_map == class_id
                color_image[mask] = enhanced_color_map[class_id]
            else:
                # 为未知类别生成随机颜色
                mask = segmentation_map == class_id
                random_color = [
                    np.random.randint(0, 256),
                    np.random.randint(0, 256),
                    np.random.randint(0, 256)
                ]
                color_image[mask] = random_color
        
        # 添加边缘增强以提高可视化效果
        gray = cv2.cvtColor(color_image, cv2.COLOR_RGB2GRAY)
        edges = cv2.Canny(gray, 50, 150)
        
        # 将边缘叠加到彩色图像上
        edge_mask = edges > 0
        color_image[edge_mask] = [255, 255, 255]  # 白色边缘
        
        return color_image
    
    def _create_vegetation_overlay(self, segmentation_map: np.ndarray, original_image: np.ndarray) -> np.ndarray:
        """
        创建植被高亮覆盖图像
        
        Args:
            segmentation_map: 分割结果图
            original_image: 原始图像
            
        Returns:
            植被高亮覆盖图像
        """
        # 创建植被掩码（类别8是vegetation）
        vegetation_mask = segmentation_map == 8
        
        # 创建高亮图像
        highlight_image = original_image.copy()
        
        # 将植被区域高亮为绿色
        highlight_image[vegetation_mask] = [0, 255, 0]  # 绿色
        
        # 将非植被区域调暗
        non_vegetation_mask = ~vegetation_mask
        highlight_image[non_vegetation_mask] = (highlight_image[non_vegetation_mask] * 0.6).astype(np.uint8)
        
        return highlight_image
    
    def clear_data(self):
        """清空数据"""
        self.results_data.clear()
        self.summary_stats.clear()
    
    def get_results_dataframe(self) -> pd.DataFrame:
        """
        获取结果DataFrame
        
        Returns:
            结果DataFrame
        """
        return pd.DataFrame(self.results_data)

# 测试函数
def test_exporter():
    """测试结果导出模块"""
    exporter = ResultExporter()
    
    # 添加测试数据
    download_result = {
        'success': True,
        'lng': 116.404,
        'lat': 39.915,
        'filepath': 'test_image.jpg',
        'download_time': '2024-01-01 12:00:00',
        'file_size': 1024000
    }
    
    analysis_result = {
        'green_view_rate': 25.5,
        'vegetation_pixels': 26214,
        'total_pixels': 102400,
        'analysis_paths': {
            'vegetation_highlight': 'test_vegetation.jpg',
            'segmentation_overlay': 'test_segmentation.jpg'
        }
    }
    
    exporter.add_result(download_result, analysis_result)
    
    # 导出测试
    exporter.export_to_excel('test_report.xlsx')
    exporter.export_to_csv('test_results.csv')
    exporter.export_summary_json('test_summary.json')
    
    print("测试完成")

if __name__ == "__main__":
    test_exporter()